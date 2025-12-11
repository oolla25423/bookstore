# -*- coding: utf-8 -*-
"""
Тесты для административной панели и функционала экспорта
Проверка прав доступа админов, экспорта в XLSX, структуры данных
"""

import io
from decimal import Decimal

from django.contrib.admin.sites import AdminSite
from django.contrib.auth import get_user_model
from django.test import RequestFactory, TestCase
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient, APITestCase

from api.admin import AuthorAdmin, BookAdmin, OrderAdmin, ReviewAdmin
from api.models import Author, Book, Order, OrderItem, Review

User = get_user_model()


class AdminAccessTestCase(TestCase):
    """Тесты доступа к административной панели"""

    def setUp(self):
        """Подготовка данных"""
        self.admin_user = User.objects.create_superuser(
            username="admin", email="admin@example.com", password="adminpass123"
        )
        self.regular_user = User.objects.create_user(
            username="user", email="user@example.com", password="userpass123"
        )
        self.client.login(username="admin", password="adminpass123")

    def test_admin_can_access_admin_panel(self):
        """Проверка что администратор имеет доступ к админ-панели"""
        response = self.client.get("/admin/")
        self.assertEqual(response.status_code, 200)

    def test_regular_user_cannot_access_admin_panel(self):
        """Проверка что обычный пользователь не имеет доступа"""
        self.client.logout()
        self.client.login(username="user", password="userpass123")

        response = self.client.get("/admin/")
        # Должен перенаправить на страницу логина
        self.assertEqual(response.status_code, 302)
        self.assertIn("/admin/login/", response.url)

    def test_guest_redirected_to_login(self):
        """Проверка что гость перенаправляется на логин"""
        self.client.logout()
        response = self.client.get("/admin/")
        self.assertEqual(response.status_code, 302)


class AdminModelDisplayTestCase(TestCase):
    """Тесты отображения моделей в админ-панели"""

    def setUp(self):
        """Подготовка данных"""
        self.site = AdminSite()
        self.admin_user = User.objects.create_superuser(
            username="admin", email="admin@example.com", password="adminpass123"
        )
        self.author = Author.objects.create(name="Test Author", bio="Test bio")
        self.book = Book.objects.create(
            title="Test Book",
            author=self.author,
            price=Decimal("500.00"),
            stock=10,
            description="Test description",
        )

    def test_book_admin_list_display(self):
        """Проверка отображения списка книг в админке"""
        book_admin = BookAdmin(Book, self.site)

        # Проверяем что есть list_display
        self.assertTrue(hasattr(book_admin, "list_display"))

        # Проверяем основные поля
        list_display = book_admin.list_display
        self.assertIn("title", list_display)
        self.assertIn("author", list_display)
        self.assertIn("price", list_display)

    def test_author_admin_list_display(self):
        """Проверка отображения списка авторов в админке"""
        author_admin = AuthorAdmin(Author, self.site)

        self.assertTrue(hasattr(author_admin, "list_display"))
        list_display = author_admin.list_display
        self.assertIn("name", list_display)

    def test_order_admin_has_inline(self):
        """Проверка что Order имеет inline для OrderItem"""
        order_admin = OrderAdmin(Order, self.site)

        # Проверяем наличие inlines
        self.assertTrue(hasattr(order_admin, "inlines"))
        self.assertTrue(len(order_admin.inlines) > 0)


class ExportXLSXTestCase(APITestCase):
    """Тесты экспорта данных в XLSX"""

    def setUp(self):
        """Подготовка данных"""
        self.client = APIClient()

        # Создаем администратора
        self.admin_user = User.objects.create_superuser(
            username="admin", email="admin@example.com", password="adminpass123", role="admin"
        )

        # Создаем обычного пользователя
        self.regular_user = User.objects.create_user(
            username="user", email="user@example.com", password="userpass123"
        )

        # Создаем тестовые данные
        self.author1 = Author.objects.create(name="Author One", bio="Bio One")
        self.author2 = Author.objects.create(name="Author Two", bio="Bio Two")

        self.book1 = Book.objects.create(
            title="Book One", author=self.author1, price=Decimal("300.00"), stock=5
        )
        self.book2 = Book.objects.create(
            title="Book Two", author=self.author2, price=Decimal("450.00"), stock=10
        )

    def test_admin_can_export_data(self):
        """Проверка что администратор может экспортировать данные"""
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/export/?model=book&fields=title&fields=price")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", response.get("Content-Disposition", ""))

    def test_regular_user_cannot_export(self):
        """Проверка что обычный пользователь не может экспортировать"""
        self.client.force_authenticate(user=self.regular_user)

        response = self.client.get("/api/export/?model=book&fields=title")

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_guest_cannot_export(self):
        """Проверка что гость не может экспортировать"""
        response = self.client.get("/api/export/?model=book&fields=title")

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_export_xlsx_structure(self):
        """Проверка структуры экспортированного XLSX файла"""
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/export/?model=book&fields=title&fields=price")

        # Проверяем что получили файл
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        # Загружаем файл из ответа
        file_content = response.content
        workbook = load_workbook(filename=io.BytesIO(file_content))

        # Проверяем наличие листов
        sheet_names = workbook.sheetnames
        self.assertIn("book", sheet_names)

    def test_export_books_sheet_data(self):
        """Проверка данных в листе Books"""
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/export/?model=book&fields=title&fields=price")

        # Загружаем файл
        file_content = response.content
        workbook = load_workbook(filename=io.BytesIO(file_content))

        # Получаем лист book
        books_sheet = workbook["book"]

        # Проверяем заголовки (первая строка)
        headers = [cell.value for cell in books_sheet[1]]
        self.assertIn("title", headers)
        self.assertIn("price", headers)

        # Проверяем количество строк (заголовок + 2 книги)
        self.assertGreaterEqual(books_sheet.max_row, 3)

    def test_export_authors_sheet_data(self):
        """Проверка данных в листе Authors"""
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/export/?model=author&fields=name")

        file_content = response.content
        workbook = load_workbook(filename=io.BytesIO(file_content))

        authors_sheet = workbook["author"]

        # Проверяем заголовки
        headers = [cell.value for cell in authors_sheet[1]]
        self.assertIn("name", headers)

        # Проверяем количество строк (заголовок + 2 автора)
        self.assertGreaterEqual(authors_sheet.max_row, 3)

    def test_export_with_table_filter(self):
        """Проверка экспорта с фильтром по таблицам"""
        self.client.force_authenticate(user=self.admin_user)

        # Экспортируем только книги
        response = self.client.get("/api/export/?model=book&fields=title")

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        file_content = response.content
        workbook = load_workbook(filename=io.BytesIO(file_content))

        # Должен быть только лист book
        sheet_names = workbook.sheetnames
        self.assertIn("book", sheet_names)

    def test_export_file_naming(self):
        """Проверка имени экспортированного файла"""
        self.client.force_authenticate(user=self.admin_user)

        response = self.client.get("/api/export/?model=book&fields=title")

        content_disposition = response["Content-Disposition"]
        self.assertIn("book.xlsx", content_disposition)


class AdminCRUDTestCase(TestCase):
    """Тесты CRUD операций через админ-панель"""

    def setUp(self):
        """Подготовка данных"""
        self.admin_user = User.objects.create_superuser(
            username="admin", email="admin@example.com", password="adminpass123"
        )
        self.client.login(username="admin", password="adminpass123")

    def test_admin_can_create_author(self):
        """Проверка создания автора через админку"""
        initial_count = Author.objects.count()

        data = {
            "name": "New Author",
            "bio": "New author bio",
        }

        response = self.client.post("/admin/api/author/add/", data)

        # Должен быть редирект после успешного создания
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Author.objects.count(), initial_count + 1)

    def test_admin_can_create_book(self):
        """Проверка создания книги через админку"""
        author = Author.objects.create(name="Test Author")
        initial_count = Book.objects.count()

        data = {
            "title": "New Book",
            "author": author.id,
            "price": "500.00",
            "stock": 10,
            "description": "Test description",
        }

        response = self.client.post("/admin/api/book/add/", data)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(Book.objects.count(), initial_count + 1)

    def test_admin_can_view_book_list(self):
        """Проверка просмотра списка книг"""
        author = Author.objects.create(name="Test Author")
        Book.objects.create(
            title="Test Book", author=author, price=Decimal("500.00"), stock=10
        )

        response = self.client.get("/admin/api/book/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Test Book")

    def test_admin_can_edit_book(self):
        """Проверка редактирования книги"""
        author = Author.objects.create(name="Test Author")
        book = Book.objects.create(
            title="Original Title", author=author, price=Decimal("500.00"), stock=10
        )

        data = {
            "title": "Updated Title",
            "author": author.id,
            "price": "600.00",
            "stock": 15,
            "description": "Updated description",
        }

        response = self.client.post(f"/admin/api/book/{book.id}/change/", data)

        book.refresh_from_db()
        self.assertEqual(book.title, "Updated Title")
        self.assertEqual(book.price, Decimal("600.00"))

    def test_admin_can_delete_book(self):
        """Проверка удаления книги"""
        author = Author.objects.create(name="Test Author")
        book = Book.objects.create(
            title="Test Book", author=author, price=Decimal("500.00"), stock=10
        )

        initial_count = Book.objects.count()

        # Подтверждение удаления
        response = self.client.post(
            f"/admin/api/book/{book.id}/delete/", {"post": "yes"}
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(Book.objects.count(), initial_count - 1)


class AdminSearchFilterTestCase(TestCase):
    """Тесты поиска и фильтрации в админ-панели"""

    def setUp(self):
        """Подготовка данных"""
        self.admin_user = User.objects.create_superuser(
            username="admin", email="admin@example.com", password="adminpass123"
        )
        self.client.login(username="admin", password="adminpass123")

        # Создаем тестовые данные
        self.author1 = Author.objects.create(name="John Smith")
        self.author2 = Author.objects.create(name="Jane Doe")

        self.book1 = Book.objects.create(
            title="Python Programming",
            author=self.author1,
            price=Decimal("500.00"),
            stock=10,
        )
        self.book2 = Book.objects.create(
            title="Django Basics", author=self.author2, price=Decimal("400.00"), stock=5
        )

    def test_admin_search_books_by_title(self):
        """Проверка поиска книг по названию"""
        response = self.client.get("/admin/api/book/", {"q": "Python"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Python Programming")
        self.assertNotContains(response, "Django Basics")

    def test_admin_search_books_by_author(self):
        """Проверка поиска книг по автору"""
        response = self.client.get("/admin/api/book/", {"q": "John"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Python Programming")


class AdminPermissionsTestCase(TestCase):
    """Тесты прав доступа в админ-панели"""

    def setUp(self):
        """Подготовка данных"""
        self.superuser = User.objects.create_superuser(
            username="super", email="super@example.com", password="superpass123"
        )

        self.staff_user = User.objects.create_user(
            username="staff",
            email="staff@example.com",
            password="staffpass123",
            is_staff=True,
        )

        self.regular_user = User.objects.create_user(
            username="regular", email="regular@example.com", password="regularpass123"
        )

    def test_superuser_has_full_access(self):
        """Проверка что суперпользователь имеет полный доступ"""
        self.client.login(username="super", password="superpass123")

        response = self.client.get("/admin/")
        self.assertEqual(response.status_code, 200)

        response = self.client.get("/admin/api/book/")
        self.assertEqual(response.status_code, 200)

    def test_staff_user_has_limited_access(self):
        """Проверка что staff пользователь имеет ограниченный доступ"""
        self.client.login(username="staff", password="staffpass123")

        # Может войти в админку
        response = self.client.get("/admin/")
        self.assertEqual(response.status_code, 200)

    def test_regular_user_no_access(self):
        """Проверка что обычный пользователь не имеет доступа"""
        self.client.login(username="regular", password="regularpass123")

        response = self.client.get("/admin/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/admin/login/", response.url)
