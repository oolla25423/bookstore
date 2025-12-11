from django.contrib.auth import authenticate
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from openpyxl import Workbook
from rest_framework import filters, generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken

from .filters import BookFilter
from .models import Author, Book, Order, OrderItem, Review, User
from .serializers import (
    AuthorSerializer,
    BookSerializer,
    LoginSerializer,
    OrderItemSerializer,
    OrderSerializer,
    RegisterSerializer,
    ReviewSerializer,
    UserSerializer,
)


# Регистрация пользователя
@swagger_auto_schema(
    method="post",
    request_body=RegisterSerializer,
    responses={201: UserSerializer(), 400: "Bad Request"},
    operation_description="Регистрация нового пользователя",
    security=[],
)
@api_view(["POST"])
@permission_classes([AllowAny])
def register_view(request):
    """
    Регистрация нового пользователя в системе.
    Возвращает данные пользователя и JWT токены.
    """
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        refresh = RefreshToken.for_user(user)
        return Response(
            {
                "user": UserSerializer(user).data,
                "refresh": str(refresh),
                "access": str(refresh.access_token),
            },
            status=status.HTTP_201_CREATED,
        )
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Авторизация пользователя
@swagger_auto_schema(
    method="post",
    request_body=LoginSerializer,
    responses={200: UserSerializer(), 400: "Bad Request"},
    operation_description="Авторизация пользователя",
    security=[],
)
@api_view(["POST"])
@permission_classes([AllowAny])
def login_view(request):
    """
    Авторизация пользователя в системе.
    Возвращает данные пользователя и JWT токены.
    """
    serializer = LoginSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.validated_data["user"]
        refresh = RefreshToken.for_user(user)
        return Response(
            {
                "user": UserSerializer(user).data,
                "refresh": str(refresh),
                "access": str(refresh.access_token),
            }
        )
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# CRUD для пользователей
class UserListCreateView(generics.ListCreateAPIView):
    """
    API для получения списка пользователей и создания нового пользователя.
    Администраторы видят всех пользователей, обычные пользователи - только себя.
    """

    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if self.request.user.role == "admin":
            return User.objects.all()
        return User.objects.filter(id=self.request.user.id)


class UserDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    API для получения, обновления и удаления пользователя.
    Администраторы могут работать с любым пользователем, обычные - только со своим профилем.
    """

    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if self.request.user.role == "admin":
            return User.objects.all()
        return User.objects.filter(id=self.request.user.id)


# CRUD для авторов
class AuthorListCreateView(generics.ListCreateAPIView):
    """
    API для получения списка авторов и создания нового автора.
    Просмотр доступен всем (включая гостей), создание - только авторизованным.
    """

    queryset = Author.objects.all()
    serializer_class = AuthorSerializer
    permission_classes = [AllowAny]  # Разрешить гостевой просмотр

    def get_permissions(self):
        if self.request.method == "POST":
            return [IsAuthenticated()]
        return [AllowAny()]


class AuthorDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    API для получения, обновления и удаления автора.
    Требуется авторизация.
    """

    queryset = Author.objects.all()
    serializer_class = AuthorSerializer
    permission_classes = [IsAuthenticated]


# CRUD для книг с фильтрацией и поиском
class BookListCreateView(generics.ListCreateAPIView):
    """
    API для получения списка книг и создания новой книги.
    Поддерживает фильтрацию по автору и цене, поиск по названию/описанию/автору, сортировку.
    Просмотр доступен всем, создание - только авторизованным.
    """

    queryset = Book.objects.all()
    serializer_class = BookSerializer
    permission_classes = [AllowAny]  # Разрешить гостевой просмотр
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = BookFilter
    search_fields = [
        "title",
        "description",
        "author__name",
    ]  # Поиск по названию, описанию, автору
    ordering_fields = ["price", "title", "created_at"]  # Сортировка

    def get_permissions(self):
        if self.request.method == "POST":
            return [IsAuthenticated()]
        return [AllowAny()]


class BookDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    API для получения, обновления и удаления книги.
    Требуется авторизация для изменения и удаления.
    """

    queryset = Book.objects.all()
    serializer_class = BookSerializer

    def get_permissions(self):
        if self.request.method == "GET":
            return [AllowAny()]
        return [IsAuthenticated()]


# CRUD для заказов
class OrderListCreateView(generics.ListCreateAPIView):
    """
    API для получения списка заказов и создания нового заказа.
    Администраторы видят все заказы, пользователи - только свои.
    """

    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["status"]
    ordering_fields = ["created_at", "total_price"]

    def get_queryset(self):
        if self.request.user.role == "admin":
            return Order.objects.all()
        return Order.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class OrderDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    API для получения, обновления и удаления заказа.
    Администраторы могут работать с любым заказом, пользователи - только со своими.
    """

    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if self.request.user.role == "admin":
            return Order.objects.all()
        return Order.objects.filter(user=self.request.user)


# Создание заказа с элементами
@swagger_auto_schema(
    method="post",
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            "items": openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "book_id": openapi.Schema(type=openapi.TYPE_INTEGER),
                        "quantity": openapi.Schema(type=openapi.TYPE_INTEGER),
                    },
                ),
            )
        },
    ),
    responses={201: OrderSerializer(), 400: "Bad Request"},
    operation_description="Создание заказа с элементами",
    security=[{"Bearer": []}],
)
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_order(request):
    """
    Создание заказа с элементами.
    Принимает список товаров (book_id, quantity), проверяет наличие на складе,
    создает заказ и элементы заказа, обновляет количество товара на складе.
    """
    items_data = request.data.get("items", [])
    if not items_data:
        return Response(
            {"error": "Не указаны товары"}, status=status.HTTP_400_BAD_REQUEST
        )

    total_price = 0
    order_items = []

    for item in items_data:
        book_id = item.get("book_id")
        quantity = item.get("quantity", 1)
        try:
            book = Book.objects.get(id=book_id)
            if book.stock < quantity:
                return Response(
                    {"error": f"Недостаточно товара {book.title}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            price = book.price * quantity
            total_price += price
            order_items.append(
                {"book": book, "quantity": quantity, "price": book.price}
            )
        except Book.DoesNotExist:
            return Response(
                {"error": f"Книга с id {book_id} не найдена"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    order = Order.objects.create(user=request.user, total_price=total_price)
    for item in order_items:
        OrderItem.objects.create(
            order=order,
            book=item["book"],
            quantity=item["quantity"],
            price=item["price"],
        )
        item["book"].stock -= item["quantity"]
        item["book"].save()

    # Calculate total price using the model method


    serializer = OrderSerializer(order)
    return Response(serializer.data, status=status.HTTP_201_CREATED)


# CRUD для отзывов
class ReviewListCreateView(generics.ListCreateAPIView):
    """
    API для получения списка отзывов и создания нового отзыва.
    Поддерживает фильтрацию по книге (параметр book).
    Просмотр доступен всем, создание - только авторизованным.
    """

    queryset = Review.objects.all()
    serializer_class = ReviewSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["book", "rating"]
    ordering_fields = ["created_at", "rating"]

    def get_permissions(self):
        if self.request.method == "POST":
            return [IsAuthenticated()]
        return [AllowAny()]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class ReviewDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    API для получения, обновления и удаления отзыва.
    Администраторы могут работать с любым отзывом, пользователи - только со своими.
    """

    queryset = Review.objects.all()
    serializer_class = ReviewSerializer

    def get_permissions(self):
        if self.request.method == "GET":
            return [AllowAny()]
        return [IsAuthenticated()]

    def get_queryset(self):
        if self.request.method == "GET":
            return Review.objects.all()
        if self.request.user.role == "admin":
            return Review.objects.all()
        return Review.objects.filter(user=self.request.user)


# Экспорт данных в XLSX для администратора
@swagger_auto_schema(
    method="get",
    manual_parameters=[
        openapi.Parameter(
            "model",
            openapi.IN_QUERY,
            description="Название модели (user, author, book, order, review)",
            type=openapi.TYPE_STRING,
        ),
        openapi.Parameter(
            "fields",
            openapi.IN_QUERY,
            description="Список полей для экспорта",
            type=openapi.TYPE_ARRAY,
            items=openapi.Items(type=openapi.TYPE_STRING),
        ),
    ],
    responses={200: "XLSX file", 400: "Bad Request", 403: "Forbidden"},
    operation_description="Экспорт данных в XLSX (только для администраторов)",
    security=[{"Bearer": []}],
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_data(request):
    """
    Экспорт данных из базы данных в формат XLSX.
    Доступен только администраторам.
    Параметры:
    - model: название модели (user, author, book, order, review)
    - fields: список полей для экспорта (можно указать несколько)
    """
    if request.user.role != "admin":
        return Response({"error": "Доступ запрещен"}, status=status.HTTP_403_FORBIDDEN)

    model_name = request.GET.get("model")
    fields = request.GET.getlist("fields")

    if not model_name or not fields:
        return Response(
            {"error": "Укажите model и fields"}, status=status.HTTP_400_BAD_REQUEST
        )

    models_dict = {
        "user": User,
        "author": Author,
        "book": Book,
        "order": Order,
        "review": Review,
        "orderitem": OrderItem,
    }

    if model_name not in models_dict:
        return Response(
            {"error": "Неверная модель"}, status=status.HTTP_400_BAD_REQUEST
        )

    model = models_dict[model_name]
    queryset = model.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = model_name

    # Заголовки
    for col_num, field in enumerate(fields, 1):
        ws.cell(row=1, column=col_num, value=field)

    # Данные
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field, "")
            ws.cell(row=row_num, column=col_num, value=str(value))

    # Сохраняем в байтовый поток
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={model_name}.xlsx"
    return response
